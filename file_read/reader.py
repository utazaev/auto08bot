import openpyxl
from aiogram import Bot, F
from aiogram.types import Message
import requests
from urllib.parse import urlencode
from loader import dp


base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
public_key = 'https://disk.yandex.ru/i/jQE7e0IKav11lQ'  # Сюда вписываете вашу ссылку

# Получаем загрузочную ссылку
final_url = base_url + urlencode(dict(public_key=public_key))
response = requests.get(final_url)
download_url = response.json()['href']

async def sort_automag(message: Message, bot: Bot): # 
    download_response = requests.get(download_url)  # скачиваем файл в директорию с проектом
    with open('Sprav08.xlsx', 'wb') as f:  # называем файл как нам нужно
        f.write(download_response.content)

    wb = openpyxl.reader.excel.load_workbook(filename='Sprav08.xlsx')  # открываем книгу эксель
    print(wb.sheetnames)  # печатаем имена листов
    wb.active = 0  # делаем активным первый лист
    sheet = wb.active  # обращаемся к листу

    for i in range(2, 100): #перебираем строки Автомагазины в первом столбце эксель файла
        if str(sheet['A' + str(i)].value) == str('Автомагазины'):
            stroka = ('<b>Магазин:</b> ' + str(sheet['B' + str(i)].value) + '\n<b>Отдел:</b> ' + str(sheet['C' + str(i)].value) + '\n<b>Телефон:</b> ' + str(sheet['D' + str(i)].value) + '\n<b>Режим работы:</b> ' + str(sheet['E' + str(i)].value) + '\n<b>Адрес:</b> ' + str(sheet['F' + str(i)].value))
            await bot.send_message(message.from_user.id, stroka, parse_mode='HTML') # соединяем все в одну строку и отправляем собщение в канал
        


async def sort_autoopen(message: Message, bot: Bot):
    download_response = requests.get(download_url)
    with open('Sprav08.xlsx', 'wb') as f: 
        f.write(download_response.content)
    
    wb = openpyxl.reader.excel.load_workbook(filename='Sprav08.xlsx')  # открываем книгу эксель
    print(wb.sheetnames)  # печатаем имена листов
    wb.active = 0  # делаем активным первый лист
    sheet = wb.active  # обращаемся к листу
    
    for i in range(2, 100):
        if str(sheet['A' + str(i)].value) == str('Аварийное вскрытие'):
            stroka_autoopen = ('<b>Имя:</b> ' + str(sheet['B' + str(i)].value) + '\n<b>Телефон:</b> ' + str(
                sheet['D' + str(i)].value) + '\n<b>Режим работы:</b> ' + str(
                sheet['E' + str(i)].value))
            await bot.send_message(message.from_user.id, stroka_autoopen, parse_mode='HTML')

async def sort_autoelectric(message: Message, bot: Bot):
    download_response = requests.get(download_url)
    with open('Sprav08.xlsx', 'wb') as f: 
        f.write(download_response.content)
    
    wb = openpyxl.reader.excel.load_workbook(filename='Sprav08.xlsx')  # открываем книгу эксель
    print(wb.sheetnames)  # печатаем имена листов
    wb.active = 0  # делаем активным первый лист
    sheet = wb.active  # обращаемся к листу
    
    for i in range(2, 100):
        if str(sheet['A' + str(i)].value) == str('Автоэлектрик'):
            stroka_autoelectric = ('<b>Имя:</b> ' + str(sheet['B' + str(i)].value) + '\n<b>Телефон:</b> ' + str(
                sheet['D' + str(i)].value) + '\n<b>Режим работы:</b> ' + str(
                sheet['E' + str(i)].value))
            await bot.send_message(message.from_user.id, stroka_autoelectric, parse_mode='HTML')